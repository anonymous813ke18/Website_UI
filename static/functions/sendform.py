import random
import re
import pandas as pd
from openpyxl import load_workbook
from static.functions import common_functions
from flask import jsonify
from datetime import datetime




def generate_form_id():
    original_id = 'abcd1234'
    id_list = list(original_id)
    random.shuffle(id_list)
    return ''.join(id_list)

def is_item_already_initiated(product_ids):

    # Read the Excel file with specific sheet name
    df = pd.read_excel('Excel/handover_data.xlsx')
    
    # Get the set of product IDs from the transaction sheet
    transaction_product_ids = set(df['ProductID'])
    
    # Find the intersection between the transaction product IDs and the provided product IDs
    common_product_ids = list(set(product_ids) & transaction_product_ids)
    
    return common_product_ids


def cart_items_function(name,project):

    # Load the Excel file
    wb = load_workbook('Excel/inventory.xlsx')
    sheet = wb.active

    # Define a list to store the dictionaries
    data = []

    # Define a list to store Serial Nos
    Serial_Nos = []

    nameproject = []

    # Insert name and project as the first dictionary in the data list
    nameproject.append({
        'Name': name,
        'Project': project
    })


    # Iterate over rows and append data to the list as dictionaries
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Check if the current owner matches the name from the session cookie
        if row[6] == name:
            # Append the Serial No to the list
            Serial_Nos.append(row[4])
            
            item = {
                'Category': row[0],
                'Name': row[1],
                'Make': row[2],
                'Model': row[3],
                'SerialNo': row[4],
                'Project' : row[5],
                'Owner' : row[6],
            }
            data.append(item)

    # Call is_item_already_initiated to receive items that are already initiated, they will be present in the transaction sheet
    already_initiated_items = is_item_already_initiated(Serial_Nos)

    # Remove items from data if their Serial No matches any in the already_initiated_items list
    data = [item for item in data if item['SerialNo'] not in already_initiated_items]

    dropdownvalues = receive_destination_dropdown_values()

    # Create a new list containing both data and result
    combined_data = [data, dropdownvalues, nameproject]
    
    jsonify(combined_data)

    return combined_data





def receive_destination_dropdown_values():
    try:

        file = 'Excel/user_info.xlsx'
        # Read the Excel file with specific sheet name
        df = pd.read_excel(file)
        
        # Filter out NaN values from 'Name' and 'Project' columns
        Names = df['Name'].dropna().tolist()  
        Projects = df['Project'].dropna().tolist()  

        # Initialize an empty dictionary
        name_project_dict = {}

        # Iterate through each name and project pair and populate the dictionary
        for name, project in zip(Names, Projects):
            name_project_dict[name] = project
        print('this is the name project dictionary',name_project_dict)
        return name_project_dict
    
    except Exception as e:
        return {'error': str(e)}


def process_form_data(form_data):
    print('we are here in the process form data functionnnnnnnnnnnnnnnnnnnn', form_data)
    try:
        # Extract form details
        form_details = form_data[0]
        source = form_details.get('Source', '')
        destination = form_details.get('Destination', '')
        sender = form_details.get('Sender', '')
        receiver = form_details.get('Receiver', '')

        # Extract item details
        item_details = form_data[1:]

        excel_data = pd.read_excel('Excel/handover_data.xlsx')
        print('this is the excel data', excel_data)
        # Generate a unique FormID
        form_ids = excel_data['FormID'].tolist()
        print('this is the formidlist already present in the excel',form_ids)
        unique_form_id = generate_form_id()
        while unique_form_id in form_ids:
            unique_form_id = generate_form_id()
        print('this is the form id that we have generated now',unique_form_id)
        # Create a DataFrame to store the new data
        df = pd.DataFrame(columns=['FormID', 'Source', 'Destination', 'Sender', 'Receiver', 'Category','Name','Make','Model','ProductID', 'SenderCondition', 'SenderRemarks', 'InitiationDate', 'Status'])
        
        # Add item details to DataFrame
        current_date_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for i, item in enumerate(item_details):
            df.loc[i+1] = [unique_form_id, source, destination, sender, receiver, item.get('Category', ''),item.get('Name', ''),item.get('Make', ''),item.get('Model', ''),item.get('ProductID', ''), item.get('SenderCondition', ''), item.get('SenderRemarks', ''), current_date_time, 'Pending']
        print('this is the df we made with the details ', df)
        # Concatenate the existing data with the new data
        updated_data = pd.concat([excel_data, df], ignore_index=True)

        # Write the updated data back to the 'transaction' sheet
        updated_data.to_excel('Excel/handover_data.xlsx', index=False)
        print('yeahhh we have done it')
        return {'message': 'Data successfully updated in the transaction sheet of the Excel file.'}
    except Exception as e:
        print('An error occurred during processing form data:', e)
        return {'error': str(e)}